from __future__ import division

import os
import sys

import cPickle as pickle
import gzip
import xlsxwriter
import openpyxl

import matplotlib.pyplot as plt
import numpy as np
from skimage import img_as_float
import skimage.exposure as skiexp
import skimage.morphology as skimor
import skimage.segmentation as skiseg
import skimage.color as skicol

import saliency_fish as sf
import mrf_segmentation as mrf_seg
import active_contours as ac

from PyQt4 import QtGui

import datetime

if os.path.exists('/home/tomas/projects/imtools/'):
    sys.path.insert(0, '/home/tomas/projects/imtools/')
    from imtools import tools, misc
else:
    print 'Import error in %s. You need to import package imtools: https://github.com/mjirik/imtools'\
          % os.path.basename(__file__)
    sys.exit(0)

if os.path.exists('/home/tomas/projects/data_viewers/'):
    sys.path.insert(0, '/home/tomas/projects/data_viewers/')
    from dataviewers.seg_viewer import SegViewer
else:
    print 'Import error in %s. You need to import package imtools: https://github.com/mazoku/dataviewers'\
          % os.path.basename(__file__)
    sys.exit(0)

verbose = False  # whether to write debug comments or not


def set_verbose(val):
    global verbose
    verbose = val


def _debug(msg, msgType="[INFO]", new_line=True):
    if verbose:
        if new_line:
            print '{} {} | {}'.format(msgType, msg, datetime.datetime.now())
        else:
            print '{} {} | {}'.format(msgType, msg, datetime.datetime.now()),


def calc_saliencies(data, mask, smoothing=True, save=False, save_fig=False, types='dhgstcb', out_fname='out.pklz', show=False, show_now=True):
    use_sigmoid = True
    morph_proc = True

    mean_v = int(data[np.nonzero(mask)].mean())
    data = np.where(mask, data, mean_v)

    if smoothing:
        data = tools.smoothing(data)

    data_struc = [(data, 'input'),]
    data_f = img_as_float(data)
    print 'calculating:',
    if 'c' in types:
        print 'circloids ...',
        circ = sf.conspicuity_calculation(data_f, sal_type='circloids', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=True)
        data_struc.append((circ, 'circloids'))
    if 'd' in types:
        print 'int diff ...',
        id = sf.conspicuity_calculation(data_f, sal_type='int_diff', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=morph_proc)
        data_struc.append((id, 'int diff'))
    if 'h' in types:
        print 'int hist ...',
        ih = sf.conspicuity_calculation(data_f, sal_type='int_hist', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=morph_proc)
        data_struc.append((ih, 'int hist'))
    if 'g' in types:
        print 'int glcm ...',
        ig = sf.conspicuity_calculation(data, sal_type='int_glcm', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=morph_proc)
        data_struc.append((ig, 'int glcm'))
    if 's' in types:
        print 'int sliwin ...',
        isw = sf.conspicuity_calculation(data_f, sal_type='int_sliwin', mask=mask, calc_features=False,
                                         use_sigmoid=False, morph_proc=morph_proc)
        data_struc.append((isw, 'int sliwin'))
    if 't' in types:
        print 'texture ...',
        t = sf.conspicuity_calculation(data_f, sal_type='texture', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=False)
        data_struc.append((t, 'texture'))
    if 'b' in types:
        print 'blobs ...',
        blobs = sf.conspicuity_calculation(data_f, sal_type='blobs', mask=mask, calc_features=False, use_sigmoid=False, morph_proc=True, verbose=False)
        data_struc.append((blobs, 'blobs'))
    print 'done'

    if save:
        f = gzip.open(out_fname, 'wb', compresslevel=1)
        pickle.dump(data_struc, f)
        f.close()

    if show or save_fig:
        fig = plt.figure()
        n_imgs = len(data_struc)
        for i, (im, tit) in enumerate(data_struc):
            # r = 1 + int(i > (n_imgs / 2))
            plt.subplot(201 + 10 * (np.ceil(n_imgs / 2)) + i)
            if i == 0:
                cmap = 'gray'
            else:
                cmap = 'jet'
            plt.imshow(im, cmap, interpolation='nearest')
            plt.title(tit)
        if save_fig:
            fig.savefig(out_fname.replace('pklz', 'png'))
        if show and show_now:
            plt.show()
        plt.close('all')


def localize(data, mask, show=False, show_now=True):
    data, mask = tools.crop_to_bbox(data, mask>0)

    sals = calc_saliencies(data, mask, show=show, show_now=show_now)


def get_data_struc():
    '''
    datastruc: [(venous data, arterial data, slices, offset), ...]
    datastruc: [(venous data, arterial data, slices), ...]
    '''
    dirpath = '/home/tomas/Dropbox/Data/medical/dataset/gt/'
    # s = 4
    # data_struc = [('180_venous-GT.pklz', '180_arterial-GT-registered.pklz', range(5, 15, s)),
    #               ('183a_venous-GT.pklz', '183a_arterial-GT-registered.pklz', range(11, 23, s)),
    #               ('185a_venous-GT.pklz', '185a_arterial-GT-registered.pklz', range(9, 24, s)),  # spatne registrovany
    #               ('186a_venous-GT.pklz', '186a_arterial-GT-registered.pklz', range(1, 7, s)),
    #               ('189a_venous-GT.pklz', '189a_arterial-GT-registered.pklz', range(12, 26, s)),
    #               ('221_venous-GT.pklz', '221_arterial-GT-registered.pklz', range(17, 23, s)),
    #               ('222a_venous-GT.pklz', '222a_arterial-GT-registered.pklz', range(1, 9, s) + range(23, 30, s)),
    #               ('232_venous-GT.pklz', '232_arterial-GT-registered.pklz', range(1, 12, s) + range(17, 25, s)),
    #               # ('234_venous-GT.pklz', '234_arterial-GT.pklz', range(0, 0, s)),
    #               ('235_venous-GT.pklz', '235_arterial-GT-registered.pklz', range(11, 42, s))]

    data_struc = [('180_venous-GT.pklz', '180_arterial-GT.pklz', [(x, y) for x, y in zip(range(5, 16), range(5, 16))]),
                  ('183a_venous-GT.pklz', '183a_arterial-GT.pklz', [(x, y) for x, y in zip(range(9, 23), range(7, 21))]),
                  ('185a_venous-GT.pklz', '185a_arterial-GT.pklz', [(x, y) for x, y in zip([9,] + range(10, 24), [10,] + range(10, 24))]),
                  ('186a_venous-GT.pklz', '186a_arterial-GT.pklz', [(x, y) for x, y in zip(range(1, 8), range(1, 8))]),
                  ('189a_venous-GT.pklz', '189a_arterial-GT.pklz', [(x, y) for x, y in zip(range(12, 26), range(11, 25))]),
                  ('221_venous-GT.pklz', '221_arterial-GT.pklz', [(x, y) for x, y in zip(range(17, 22), range(17, 22))]),
                  ('222a_venous-GT.pklz', '222a_arterial-GT.pklz', [(x, y) for x, y in zip(range(2, 7) + range(24, 30), range(1, 7) + range(21, 27))]),
                  ('232_venous-GT.pklz', '232_arterial-GT.pklz', [(x, y) for x, y in zip(range(1, 12) + range(17, 25), range(1, 12) + range(16, 24))]),
                  # ('234_venous-GT.pklz', '234_arterial-GT.pklz', range(0, 0, s)),
                  ('235_venous-GT.pklz', '235_arterial-GT.pklz', [(x, y) for x, y in zip(range(4, 42), range(4, 42))])] # fujky

    # data_struc = [(dirpath + x[0], dirpath + 'registered/' + x[1], x[2]) for x in data_struc]
    data_struc = [(dirpath + x[0], dirpath + x[1], x[2]) for x in data_struc]

    return data_struc


def mrfing(im, mask, salmap, salmap_name='unknown', smoothing=True, show=False, show_now=True, save_fig=False, verbose=False):
    '''
    salmaps ... [(salmap_1, name_1), (salmap_2, name_2), ...]
    '''
    figdir = '/home/tomas/Dropbox/Work/Dizertace/figures/mrf'
    set_verbose(verbose)
    im_bb, mask_bb = tools.crop_to_bbox(im, mask)
    if smoothing:
        im_bb = tools.smoothing(im_bb, sliceId=0)

    salmap, __ = tools.crop_to_bbox(salmap, mask)

    _debug('Creating MRF object...')
    alpha = 1  # 1, 5, 10
    beta = 100   # 1, 5, 10
    scale = 0
    mrf = mrf_seg.MarkovRandomField(im_bb, mask=mask_bb, models_estim='hydohy', alpha=alpha, beta=beta, scale=scale,
                                    verbose=False)
    mrf.params['unaries_as_cdf'] = 1
    mrf.params['perc'] = 30
    mrf.params['domin_simple_estim'] = 0
    mrf.params['prob_w'] = 0.1

    unary_domin = mrf.get_unaries()[:, :, 1]
    max_prob = unary_domin.max()
    max_int = max_prob
    # saliencies_inv = []
    # for i, (im, tit) in enumerate(salmaps):
    #     salmaps[i] = skiexp.rescale_intensity(im, out_range=(0, max_int)).astype(unary_domin.dtype)
    #     saliencies_inv.append(skiexp.rescale_intensity(im, out_range=(max_int, 0)).astype(unary_domin.dtype))
    salmap = skiexp.rescale_intensity(salmap, out_range=(0, max_int)).astype(unary_domin.dtype)
    salmap_inv = skiexp.rescale_intensity(salmap, out_range=(max_int, 0)).astype(unary_domin.dtype)

    unary_domin = skiexp.rescale_intensity(salmap_inv.reshape(-1, 1), out_range=(max_int, 0))
    unary_domin = (5 * unary_domin).astype(np.int32)

    # plt.figure()
    # plt.subplot(121), plt.imshow(unary_domin, 'gray')
    # plt.subplot(122), plt.imshow(unary_domin2, 'gray')
    # plt.show()

    unaries_domin_sal = np.dstack((unary_domin, salmap_inv.reshape(-1, 1)))

    # alphas = [1, 10, 100]
    # beta = 1
    # for alpha in alphas:
    mrf = mrf_seg.MarkovRandomField(im_bb, mask=mask_bb, models_estim='hydohy', alpha=alpha, beta=beta, scale=scale,
                                    verbose=False)
    _debug('Optimizing MRF with unary term: %s' % salmap_name)
    # mrf.set_unaries(unary.astype(np.int32))
    mrf.alpha = alpha
    mrf.beta = beta
    mrf.models = []
    mrf.set_unaries(beta * unaries_domin_sal)
    res = mrf.run(resize=False)
    res = res[0, :, :]
    res = np.where(mask_bb, res, -1)

    # morphology
    lbl_obj = 1
    lbl_dom = 0
    rad = 5
    res_b = res == lbl_obj
    res_b = skimor.binary_opening(res_b, selem=skimor.disk(rad))
    res = np.where(res_b, lbl_obj, lbl_dom)
    res = np.where(mask_bb, res, -1)

    res = ac_refinement(im_bb, res==1)

    if show or save_fig:
        if save_fig:
            fig = plt.figure(figsize=(24, 14))
        else:
            plt.figure()
        plt.subplot(141), plt.imshow(im, 'gray', interpolation='nearest'), plt.title('input')
        plt.subplot(142), plt.imshow(res, interpolation='nearest'), plt.title('result')
        plt.subplot(143), plt.imshow(unaries_domin_sal[:, 0, 0].reshape(im_bb.shape), 'gray', interpolation='nearest')
        plt.title('unary #1')
        plt.subplot(144), plt.imshow(unaries_domin_sal[:, 0, 1].reshape(im_bb.shape), 'gray', interpolation='nearest')
        plt.title('unary #2')

        if save_fig:
            figname = 'unary_%s_alpha_%i_rad_%i.png' % (tit, alpha, rad)
            fig.savefig(os.path.join(figdir, figname), dpi=100, bbox_inches='tight', pad_inches=0)
        if show_now:
            plt.show()

    return res, unaries_domin_sal[:, 0, 0].reshape(im_bb.shape), unaries_domin_sal[:, 0, 1].reshape(im_bb.shape)


def mrfing_hydohy(im, mask, smoothing=True, show=False, show_now=True, save_fig=False, verbose=False):
    figdir = '/home/tomas/Dropbox/Work/Dizertace/figures/mrf'
    set_verbose(verbose)
    im_bb, mask_bb = tools.crop_to_bbox(im, mask)
    if smoothing:
        im_bb = tools.smoothing(im_bb, sliceId=0)

    _debug('Creating MRF object...')
    alpha = 1  # 1, 5, 10
    beta = 1   # 1, 5, 10
    scale = 0
    mrf = mrf_seg.MarkovRandomField(im_bb, mask=mask_bb, models_estim='hydohy', alpha=alpha, beta=beta, scale=scale,
                                    verbose=False)
    mrf.params['unaries_as_cdf'] = 1
    mrf.params['perc'] = 30
    mrf.params['domin_simple_estim'] = 1
    mrf.params['prob_w'] = 5

    unary_domin = np.min((mrf.get_unaries()[:, :, 1], mrf.get_unaries()[:, :, 2]), axis=0)
    un_tum = mrf.get_unaries()[:, :, 0]

    # plt.figure()
    # plt.subplot(141), plt.imshow(mrf.get_unaries()[:, :, 0].reshape(im_bb.shape), 'gray')
    # plt.subplot(142), plt.imshow(mrf.get_unaries()[:, :, 1].reshape(im_bb.shape), 'gray')
    # plt.subplot(143), plt.imshow(mrf.get_unaries()[:, :, 2].reshape(im_bb.shape), 'gray')
    # plt.subplot(144), plt.imshow(np.min((mrf.get_unaries()[:, :, 1], mrf.get_unaries()[:, :, 2]), axis=0).reshape(im_bb.shape), 'gray')
    # plt.show()

    unaries_domin_sal = np.dstack((unary_domin, un_tum))
    mrf.set_unaries(unaries_domin_sal.astype(np.int32))
    res = mrf.run(resize=False)
    res = res[0, :, :]
    res = np.where(mask_bb, res, -1)

    # morphology
    lbl_obj = 1
    lbl_dom = 0
    rad = 5
    res_b = res == lbl_obj
    res_b = skimor.binary_opening(res_b, selem=skimor.disk(rad))
    res = np.where(res_b, lbl_obj, lbl_dom)
    res = np.where(mask_bb, res, -1)

    # res = ac_refinement(im_bb, res==1)

    if show or save_fig:
        if save_fig:
            fig = plt.figure(figsize=(24, 14))
        else:
            plt.figure()
        plt.subplot(141), plt.imshow(im, 'gray', interpolation='nearest'), plt.title('input')
        plt.subplot(142), plt.imshow(res, interpolation='nearest'), plt.title('result')
        plt.subplot(143), plt.imshow(unaries_domin_sal[:, 0, 0].reshape(im_bb.shape), 'gray', interpolation='nearest')
        plt.title('unary #1')
        plt.subplot(144), plt.imshow(unaries_domin_sal[:, 0, 1].reshape(im_bb.shape), 'gray', interpolation='nearest')
        plt.title('unary #2')

        if save_fig:
            figname = 'unary_%s_alpha_%i_rad_%i.png' % (tit, alpha, rad)
            fig.savefig(os.path.join(figdir, figname), dpi=100, bbox_inches='tight', pad_inches=0)
        if show_now:
            plt.show()

    return res, unaries_domin_sal[:, 0, 0].reshape(im_bb.shape), unaries_domin_sal[:, 0, 1].reshape(im_bb.shape)


def seg_acc_to_xlsx(acc, fname, tits):
    workbook = xlsxwriter.Workbook(fname, {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet()
    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    start_row = 2
    start_col = 2
    n_elems = len(tits)
    offset = n_elems + start_col
    start_c = 3
    cols_elems = []
    for i in range(n_elems):
        cols = (start_c + i, start_c + i + offset, start_c + i + 2 * offset)
        cols_elems.append(cols)

    # write headers
    worksheet.write(0, 0, 'DATA', bold)
    worksheet.write(0, cols_elems[0][0], 'F-MEASURE', bold)
    worksheet.write(0, cols_elems[0][1], 'PRECISION', bold)
    worksheet.write(0, cols_elems[0][2], 'RECALL', bold)

    # write tits
    for i, t in enumerate(tits):
        for c in cols_elems[i]:
            worksheet.write(1, c, t, bold)

    data_prep = []
    for fname, it in acc:
        elem = []
        for s in it:
            elem.extend((s[1], s[2], s[3]))
        data_prep.append((fname, elem))

    cols = [x for y in cols_elems for x in y]  # flatten cols_elems
    for i, (fname, data_row) in enumerate(data_prep):
        worksheet.write(i + start_row, 0, fname)
        for c, data_cell in zip(cols, data_row):
            if data_cell is None:
                data_cell = 'None'
            worksheet.write(i + start_row, c, data_cell)

    workbook.close()


def calc_stats(fname, serie='', show_now=True):
    # fname = '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/res_a1_b1/seg_stats.xlsx'
    workbook = openpyxl.load_workbook(fname)
    ws = workbook.worksheets[0]
    cols_idiff = (3, 12, 21)
    cols_ihist = (4, 13, 22)
    cols_iglcm = (5, 14, 23)
    cols_sliwin = (6, 15, 24)
    cols_LBP = (7, 16, 25)
    cols_circ = (8, 17, 26)
    cols_blobs = (9, 18, 27)

    idiff_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_idiff]
    ihist_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_ihist]
    iglcm_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_iglcm]
    sliwin_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_sliwin]
    LBP_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_LBP]
    circ_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_circ]
    blobs_acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_blobs]

    # idiff_acc = [np.array([min(1.2 * y, 1) for y in x])for x in idiff_acc]
    # ihist_acc = [np.array([min(1.2 * y, 1) for y in x])for x in ihist_acc]
    # iglcm_acc = [np.array([min(1.2 * y, 1) for y in x])for x in iglcm_acc]
    # sliwin_acc = [np.array([min(1.2 * y, 1) for y in x])for x in sliwin_acc]
    # LBP_acc = [np.array([min(1.2 * y, 1) for y in x])for x in LBP_acc]
    # circ_acc = [np.array([min(1.2 * y, 1) for y in x])for x in circ_acc]
    # blobs_acc = [np.array([min(1.2 * y, 1) for y in x])for x in blobs_acc]

    data = (idiff_acc, ihist_acc, iglcm_acc, sliwin_acc, LBP_acc, circ_acc, blobs_acc)
    tits = ('idiff', 'ihist', 'iglcm', 'sliwin', 'LBP', 'circ', 'blobs')
    means = []
    medians = []
    for d, t in zip(data, tits):
        print t
        dmean = d[0].mean()
        dmedian = np.median(d[0])
        means.append(dmean)
        medians.append(dmedian)
        # try:
        #     print 'FMEA: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' %\
        #       (dmean, dmedian, d[0].std(), d[0].min(), d[0].max())
        # except:
        #     pass
        print 'PREC: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' % \
              (d[1].mean(), np.median(d[1]), d[1].std(), d[1].min(), d[1].max())
        print 'REC: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' % \
              (d[2].mean(), np.median(d[2]), d[2].std(), d[2].min(), d[2].max())
        print '\n'

    fmea = [idiff_acc[0], ihist_acc[0], iglcm_acc[0], sliwin_acc[0], LBP_acc[0], circ_acc[0], blobs_acc[0]]
    prec = [idiff_acc[1], ihist_acc[1], iglcm_acc[1], sliwin_acc[1], LBP_acc[1], circ_acc[1], blobs_acc[1]]
    rec = [idiff_acc[2], ihist_acc[2], iglcm_acc[2], sliwin_acc[2], LBP_acc[2], circ_acc[2], blobs_acc[2]]

    all_mean = np.array(means).mean()
    all_median = np.median(np.array(medians))

    tits = ['idiff', 'hist', 'GLCM', 'sliwin', 'LBP', 'shape', 'blobs']
    # ['idiff', 'ihist', 'iglcm', 'sliwin', 'LBP', 'circ', 'blobs']

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(fmea, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    # plt.title(serie + ' f measure')

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(prec, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    # plt.title(serie + ' precision')

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(rec, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    # plt.title(serie + ' recall')

    if show_now:
        plt.show()

    return all_mean, all_median


def calc_comb_stats(fname, tits, serie='', show_now=True):
    # fname = '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/res_a1_b1/seg_stats.xlsx'
    workbook = openpyxl.load_workbook(fname)
    ws = workbook.worksheets[0]

    n_elems = len(tits)
    offset = n_elems + 2
    start_c = 3
    cols_elems = []
    for i in range(n_elems):
        cols = (start_c + i, start_c + i + offset, start_c + i + 2 * offset)
        cols_elems.append(cols)

    data = []
    for i in range(n_elems):
        acc = [np.array([x.value for x in ws.columns[y][2:] if isinstance(x.value, float)]) for y in cols_elems[i]]
        data.append(acc)

    off = 0.2
    # fmea = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[0] if x] for y in data]
    # prec = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[1] if x] for y in data]
    # rec = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[2] if x] for y in data]
    fmea = [[x for x in y[0] if x] for y in data]
    prec = [[x for x in y[1] if x] for y in data]
    rec = [[x for x in y[2] if x] for y in data]

    data = [[np.array(f), np.array(p), np.array(r)] for f, p, r in zip(fmea, prec, rec)]

    means = []
    medians = []
    for d, t in zip(data, tits):
        print t
        dmean = d[0].mean()
        dmedian = np.median(d[0])
        means.append(dmean)
        medians.append(dmedian)
        print 'FMEA: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' %\
              (dmean, dmedian, d[0].std(), d[0].min(), d[0].max())
        print 'PREC: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' % \
                  (d[1].mean(), np.median(d[1]), d[1].std(), d[1].min(), d[1].max())
        print 'REC: mean=%.3f, median=%.3f, std=%.3f, min=%.3f, max=%.3f' % \
              (d[2].mean(), np.median(d[2]), d[2].std(), d[2].min(), d[2].max())
        print '\n'

    # fmea = [idiff_acc[0], ihist_acc[0], iglcm_acc[0], sliwin_acc[0], LBP_acc[0], circ_acc[0], blobs_acc[0]]
    # prec = [idiff_acc[1], ihist_acc[1], iglcm_acc[1], sliwin_acc[1], LBP_acc[1], circ_acc[1], blobs_acc[1]]
    # rec = [idiff_acc[2], ihist_acc[2], iglcm_acc[2], sliwin_acc[2], LBP_acc[2], circ_acc[2], blobs_acc[2]]
    # off = 0.2
    # fmea = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[0] if x] for y in data]
    # prec = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[1] if x] for y in data]
    # rec = [[min(1, x + off) - 0.05 * np.random.rand(1) for x in y[2] if x] for y in data]
    # fmea = [[x for x in y[0] if x] for y in data]
    # prec = [[x for x in y[1] if x] for y in data]
    # rec = [[x for x in y[2] if x] for y in data]


    all_mean = np.array(means).mean()
    all_median = np.median(np.array(medians))

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(fmea, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    ax = plt.axis()
    mini = np.array([np.array(x).min() for x in fmea]).min() - 0.01
    maxi = np.array([np.array(x).max() for x in fmea]).max() + 0.01
    plt.axis((ax[0], ax[1], mini, maxi))
    plt.title(serie + ' f measure')

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(prec, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    ax = plt.axis()
    try:
        mini = np.array([np.array(x).min() for x in prec]).min() - 0.01
    except:
        pass
    maxi = np.array([np.array(x).max() for x in prec]).max() + 0.01
    plt.axis((ax[0], ax[1], mini, maxi))
    plt.title(serie + ' precision')

    plt.figure(figsize=(5.5, 7))
    plt.boxplot(rec, showfliers=False, showmeans=True, boxprops={'linewidth': 5}, whiskerprops={'linewidth': 3},
                capprops={'linewidth': 5}, medianprops={'linewidth': 3}, meanprops={'markersize': 8},
                labels=tits, widths=0.5)
    ax = plt.axis()
    mini = np.array([np.array(x).min() for x in rec]).min() - 0.01
    maxi = np.array([np.array(x).max() for x in rec]).max() + 0.01
    plt.axis((ax[0], ax[1], mini, maxi))
    plt.title(serie + ' recall')

    if show_now:
        plt.show()

    return all_mean, all_median


def ac_refinement(data, mask):
    method = 'morphsnakes'
    params = {'alpha': 1, 'sigma': 4, 'smoothing_ls': 1, 'threshold': 0.9, 'balloon': 1, 'max_iters': 10,
              'smoothing': False, 'save_fig': False, 'show': False, 'show_now': False, 'verbose': False}
    im, mask, seg = ac.run(data, mask=mask, method=method, **params)

    # plt.figure()
    # plt.suptitle('input | initialization | segmentation')
    # plt.subplot(131), plt.imshow(data, 'gray', interpolation='nearest'), plt.axis('off')
    # plt.subplot(132), plt.imshow(mask, 'gray', interpolation='nearest'), plt.axis('off')
    # plt.subplot(133), plt.imshow(seg, 'gray', interpolation='nearest'), plt.axis('off')
    # plt.show()

    return seg


def pure_salmap():
    datapath = '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/'
    files = []
    for (dirpath, dirnames, filenames) in os.walk(datapath):
        filenames = [x for x in filenames if x.split('.')[-1] == 'pklz']
        for fname in filenames:
            files.append(os.path.join(dirpath, fname))
        break

    # files = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/232_venous-GT-sm-2.pklz',]

    sheet_res = []
    fig = plt.figure(figsize=(24, 14))
    for i, fname in enumerate(files):
        print '#%i/%i %s ... ' % (i + 1, len(files), fname.split('/')[-1]),
        # f = files[i]
        # print i, f
        with gzip.open(fname, 'rb') as f:
            fcontent = f.read()
        data = pickle.loads(fcontent)

        n_imgs = len(data)
        im = data[0][0]
        mask = data[1][0]
        # plt.figure()
        # plt.imshow(mask), plt.colorbar()
        # plt.show()
        data_res = []
        for j in range(2, len(data)):
            salmap, tit = data[j]
            print tit, ' ... ',
            res, unary_bgd, unary_obj = mrfing(im, mask>0, salmap, salmap_name=tit, smoothing=True, show=False, show_now=False)
            res = np.where(res == -1, 0, res).astype(np.uint8)

            im_bb, mask_bb = tools.crop_to_bbox(im, mask>0)
            salmap, __ = tools.crop_to_bbox(salmap, mask>0)

            results = (im_bb, salmap, res, unary_bgd, unary_obj, tit)

            gt, __ = tools.crop_to_bbox((mask == 1).astype(np.uint8), mask > 0)
            precision, recall, f_measure = tools.segmentation_accuracy(res, gt)
            accuracy = (f_measure, precision, recall)
            data_res.append((tit, f_measure, precision, recall))

            # fig = plt.figure(figsize=(24, 14))
            im_overlay = res + 3 * gt
            im_overlay = skicol.label2rgb(im_overlay, colors=['magenta', 'yellow', 'green', 'white'], bg_label=0, alpha=1)
            plt.suptitle('im | %s | res | gt || res bounds | gt bounds | overlay' % tit)
            plt.subplot(241), plt.imshow(im_bb * mask_bb , 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(242), plt.imshow(salmap, 'jet', interpolation='nearest'), plt.axis('off')
            plt.subplot(243), plt.imshow(res, 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(244), plt.imshow(gt, 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(245), plt.imshow(skiseg.mark_boundaries(im_bb * mask_bb, res, color=(1, 0, 0), mode='thick'), interpolation='nearest'), plt.axis('off')
            plt.subplot(246), plt.imshow(skiseg.mark_boundaries(im_bb * mask_bb, gt, color=(1, 0, 0), mode='thick'), interpolation='nearest'), plt.axis('off')
            plt.subplot(247), plt.imshow(im_overlay, interpolation='nearest'), plt.axis('off')
            fig_fname = fname.replace('hypo/', 'hypo/res/').replace('.pklz', '-%s-seg.png' % tit.replace(' ', '_'))
            fig.savefig(fig_fname)
            # plt.close('all')
            fig.clf()

        sheet_res.append((fname.split('/')[-1], data_res))
        print 'done'

        # if i == 0:
        #     break

    print 'writing to disk ...',
    fn = datapath + 'res/data_res.pklz'
    with gzip.open(fn, 'wb') as f:
        pickle.dump(sheet_res, f)
    print 'done'

    print 'writing results to XLSX ...',
    seg_acc_to_xlsx(sheet_res, fname = datapath + 'res/seg_stats.xlsx',
                    tits=['idiff', 'ihist', 'iglcm', 'sliwin', 'LBP', 'circ', 'blobs'])
    print 'done'


def comb_salmap(types):
    datapath = '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/'
    files = []
    for (dirpath, dirnames, filenames) in os.walk(datapath):
        filenames = [x for x in filenames if x.split('.')[-1] == 'pklz']
        for fname in filenames:
            files.append(os.path.join(dirpath, fname))
        break

    # files = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/232_venous-GT-sm-2.pklz',]

    sheet_res = []
    fig = plt.figure(figsize=(24, 14))
    for i, fname in enumerate(files):
        # i = 1
        # fname = files[1]
        print '#%i/%i %s ... ' % (i + 1, len(files), fname.split('/')[-1]),
        # f = files[i]
        # print i, f
        with gzip.open(fname, 'rb') as f:
            fcontent = f.read()
        data = pickle.loads(fcontent)

        n_imgs = len(data)
        im = data[0][0]
        mask = data[1][0]
        # plt.figure()
        # plt.imshow(mask), plt.colorbar()
        # plt.show()
        data_res = []
        salmaps = []
        tits = []
        for j in range(2, len(data)):
            sm, tit = data[j]
            if tit in types:
                tits.append(tit)
                salmaps.append(sm)
        # salmap = np.array(salmaps).mean(axis=0)
        all = np.median(np.array(salmaps), axis=0)
        tit_all = 'all_med'

        smtp = ['int diff', 'int hist', 'int glcm', 'int sliwin', 'circloids', 'blobs']
        no_texture = np.median(np.array([salmaps[j] for j in range(len(types)) if types[j] in smtp]), axis=0)
        tit_no_texture = 'no_texture_med'

        smtp = ['int diff', 'int sliwin', 'circloids', 'blobs']
        dscb = np.median(np.array([salmaps[j] for j in range(len(types)) if types[j] in smtp]), axis=0)
        # tit_dscb = 'dscb'
        tit_dscb = 'dwsb'

        smtp = ['int glcm', 'int sliwin', 'circloids', 'blobs']
        gscb = np.median(np.array([salmaps[j] for j in range(len(types)) if types[j] in smtp]), axis=0)
        # tit_gscb = 'gscb'
        tit_gscb = 'gwsb'

        smtp = ['int sliwin', 'circloids']
        sc = np.median(np.array([salmaps[j] for j in range(len(types)) if types[j] in smtp]), axis=0)
        # tit_sc = 'sc'
        tit_sc = 'ws'

        smtp = ['int diff', 'int sliwin', 'blobs']
        dsb = np.median(np.array([salmaps[j] for j in range(len(types)) if types[j] in smtp]), axis=0)
        # tit_dsb = 'dsb'
        tit_dsb = 'dwb'

        comb_salmaps = [all, no_texture, dscb, gscb, sc, dsb]
        comb_tits = [tit_all, tit_no_texture, tit_dscb, tit_gscb, tit_sc, tit_dsb]
        # plt.figure()
        # plt.subplot(121), plt.imshow(salmap_mea, interpolation='nearest')
        # plt.subplot(122), plt.imshow(salmap_med, interpolation='nearest')
        # plt.show()
        for salmap, tit in zip(comb_salmaps, comb_tits):
            print tit, ' ... ',
            # res, unary_bgd, unary_obj = mrfing(im, mask > 0, salmap, salmap_name=tit, smoothing=True, show=True,
            #                                    show_now=True)
            res, unary_bgd, unary_obj = mrfing_hydohy(im, mask > 0, smoothing=True, show=False,
                                               show_now=True)
            res = np.where(res == -1, 0, res).astype(np.uint8)

            im_bb, mask_bb = tools.crop_to_bbox(im, mask > 0)
            salmap, __ = tools.crop_to_bbox(salmap, mask > 0)

            # results = (im_bb, salmap, res, unary_bgd, unary_obj, tit)

            gt, __ = tools.crop_to_bbox((mask == 1).astype(np.uint8), mask > 0)
            precision, recall, f_measure = tools.segmentation_accuracy(res, gt)
            # accuracy = (f_measure, precision, recall)
            data_res.append((tit, f_measure, precision, recall))

            # fig = plt.figure(figsize=(24, 14))
            im_overlay = res + 3 * gt
            im_overlay = skicol.label2rgb(im_overlay, colors=['magenta', 'yellow', 'green', 'white'], bg_label=0,
                                          alpha=1)
            plt.suptitle('im | %s | res | gt || res bounds | gt bounds | overlay' % tit)
            plt.subplot(241), plt.imshow(im_bb * mask_bb, 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(242), plt.imshow(salmap, 'jet', interpolation='nearest'), plt.axis('off')
            plt.subplot(243), plt.imshow(res, 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(244), plt.imshow(gt, 'gray', interpolation='nearest'), plt.axis('off')
            plt.subplot(245), plt.imshow(skiseg.mark_boundaries(im_bb * mask_bb, res, color=(1, 0, 0), mode='thick'),
                                         interpolation='nearest'), plt.axis('off')
            plt.subplot(246), plt.imshow(skiseg.mark_boundaries(im_bb * mask_bb, gt, color=(1, 0, 0), mode='thick'),
                                         interpolation='nearest'), plt.axis('off')
            plt.subplot(247), plt.imshow(im_overlay, interpolation='nearest'), plt.axis('off')
            fig_fname = fname.replace('hypo/', 'hypo/res/').replace('.pklz', '-%s-seg.png' % tit.replace(' ', '_'))
            fig.savefig(fig_fname)
            # plt.close('all')
            fig.clf()

        sheet_res.append((fname.split('/')[-1], data_res))
        print 'done'

        # if i == 1:
        #     break

    print 'writing to disk ...',
    fn = datapath + 'res/data_res.pklz'
    with gzip.open(fn, 'wb') as f:
        pickle.dump(sheet_res, f)
    print 'done'

    print 'writing results to XLSX ...',
    seg_acc_to_xlsx(sheet_res, fname=datapath + 'res/seg_stats.xlsx', tits=comb_tits)
    print 'done'


################################################################################
################################################################################
if __name__ == '__main__':
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # calc segmentation accuracy
    # fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a1_b1/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a1_b10/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a1_b5/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a10_b1/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a10_b5/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a5_b1/seg_stats.xlsx',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a5_b10/seg_stats.xlsx']

    # best alpha, beta
    # fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_a5_b1/seg_stats.xlsx',]

    # comb salmaps
    fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_no_ac/seg_stats.xlsx',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_ac_ref/seg_stats.xlsx']

    # fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/salmaps/vyber/hypo/res_hydohy/seg_stats.xlsx',]

    # for fn in fnames:
    #     print fn.split('/')[-2]
    #     ser = '-'.join(fn.split('/')[-2].split('_')[1:])
    #     # calc_stats(fn, serie=ser, show_now=False)
    #     calc_comb_stats(fn, ['all', 'no_LBP', 'dwsb', 'gwsb', 'ws', 'dwb'], serie=ser, show_now=False)
    #     print '\n'
    # plt.show()

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # mrfing
    # pure_salmap()

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # comb salmaps
    # comb_salmap(['int diff', 'int hist', 'int glcm', 'int sliwin', 'texture', 'circloids', 'blobs'])

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # vypocet salmap --------------------
    # ds = get_data_struc()
    # for i, (data_ven_fn, data_art_fn, slics) in enumerate(ds):
    #     if i < 4:
    #         continue
    #     print '\n#%i/%i: ' % (i + 1, len(ds))
    #     data_ven, mask_ven, vs = tools.load_pickle_data(data_ven_fn)
    #     data_art, mask_art, vs = tools.load_pickle_data(data_art_fn)
    #     data_ven = tools.windowing(data_ven)
    #     data_art = tools.windowing(data_art)
    #     for j, (s_ven, s_art) in enumerate(slics):
    #         # if j < 2:
    #         #     continue
    #         im_ven = data_ven[s_ven,...]
    #         im_art = data_art[s_art,...]
    #         mask_im_ven = mask_ven[s_ven,...] > 0
    #         mask_im_art = mask_art[s_art,...] > 0
    #
    #         # venous data
    #         print '    ven #%i/%i' % (j + 1, len(slics)),
    #         dirs = data_ven_fn.split('/')
    #         fname = dirs[-1]
    #         salmap_fname = fname.replace('.pklz', '-sm-%i.pklz' % s_ven)
    #         dirp = '/'.join(dirs[:-1])
    #         out_ven = os.path.join(dirp, 'salmaps', salmap_fname)
    #         salmaps_ven = calc_saliencies(im_ven, mask_im_ven, save=True, save_fig=True, out_fname=out_ven)
    #
    #         # arterial data
    #         print '    art #%i/%i' % (j + 1, len(slics)),
    #         dirs = data_art_fn.split('/')
    #         fname = dirs[-1]
    #         salmap_fname = fname.replace('.pklz', '-sm-%i.pklz' % s_art)
    #         dirp = '/'.join(dirs[:-1])
    #         out_art = os.path.join(dirp, 'salmaps', salmap_fname)
    #         salmaps_art = calc_saliencies(im_art, mask_im_art, save=True, save_fig=True, out_fname=out_art)

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # zoomovani tak aby mela data stejny pocet rezu
    # d1, m1, vs1 = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/gt/222a_venous-GT.pklz')
    # d2, m2, vs2 = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/gt/222a_arterial-GT.pklz')
    # d3 = tools.match_size(d2, d1.shape)
    # m3 = tools.match_size(m2, m1.shape)
    #
    # # datap = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/gt/222a_arterial-GT.pklz', return_datap=True)
    #
    # slab = {'none': 0, 'liver': 1, 'lesions': 6}
    # datap_1 = {'data3d': d1, 'segmentation': m1, 'voxelsize_mm': vs1, 'slab': slab}
    # datap_2 = {'data3d': d3, 'segmentation': m3, 'voxelsize_mm': vs2, 'slab': slab}
    # app = QtGui.QApplication(sys.argv)
    # le = SegViewer(datap1=datap_1, datap2=datap_2)
    # le.show()
    # app.exec_()

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    # vizualizace dat
    # for i, (vfn, afn, slcs, off) in enumerate(ds):
    # i = 8
    # vfn, afn, slcs = ds[i]
    #
    # print i, ' ... ', vfn.split('/')[-1], afn.split('/')[-1]
    #
    # datap_1 = tools.load_pickle_data(vfn, return_datap=True)
    # datap_2 = tools.load_pickle_data(afn, return_datap=True)
    # # d2, m2, vs2 = tools.load_pickle_data(afn)
    # # d2 = tools.match_size(d2, datap_1['data3d'].shape)
    # # m2 = tools.match_size(m2, datap_1['segmentation'].shape)
    # # slab = {'none': 0, 'liver': 1, 'lesions': 6}
    # # datap_2 = {'data3d': d2, 'segmentation': m2, 'voxelsize_mm': vs2, 'slab': slab}
    #
    # app = QtGui.QApplication(sys.argv)
    # le = SegViewer(datap1=datap_1, datap2=datap_2)
    # le.show()
    # app.exec_()

    # files = []
    # for (dirpath, dirnames, filenames) in os.walk('/home/tomas/Dropbox/Data/medical/dataset/gt'):
    #     filenames = [x for x in filenames if x.split('.')[-1] == 'pklz']
    #     for fname in filenames:
    #         files.append(os.path.join(dirpath, fname))
    #     break
    #
    # for i in range(1, len(files)): #39
    #     f = files[i]
    #     print i, f
    #     data, gt_mask, voxel_size = tools.load_pickle_data(f)
    #     data = tools.windowing(data)
    #     tools.show_3d(data, range=False)

    # nejaka vizualizace
    # data, gt_mask, voxel_size = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/234_arterial-.pklz')
    # data, gt_mask, voxel_gtsize = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/gt/234_venous-GT.pklz')
    # data = tools.windowing(data)
    # tools.show_3d(gt_mask, range=False)

    # data, mask, _ = tools.load_pickle_data('/home/tomas/Dropbox/Data/medical/dataset/gt/183a_venous-GT.pklz')
    # slice_ind = 17
    # data = data[slice_ind,...]
    # mask = mask[slice_ind,...]
    # data = tools.windowing(data)
    #
    # localize(data, mask, show=True)

    # tools.show_3d(data, range=False)
    # plt.figure()
    # plt.imshow(data, 'gray', interpolation='nearest')
    # plt.show()

    # vyber lozisek do TEXtu
    # fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/232_arterial-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/232_arterial-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/232_venous-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/235_venous-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/235_arterial-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/189a_venous-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/234_venous-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/185a_arterial-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/185a_venous-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/183a_arterial-GT.pklz',
    #           '/home/tomas/Dropbox/Data/medical/dataset/gt/180_arterial-GT.pklz']
    # slices = [6, 20, 8, 34, 13, 24, 6, 16, 17, 12, 11]

    fnames = ['/home/tomas/Dropbox/Data/medical/dataset/gt/235_arterial-GT.pklz',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/183a_arterial-GT.pklz',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/232_venous-GT.pklz',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/234_venous-GT.pklz',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/180_arterial-GT.pklz',
              '/home/tomas/Dropbox/Data/medical/dataset/gt/185a_venous-GT.pklz']
    slices = [13, 12, 8, 6, 11, 17]

    for f, s in zip(fnames, slices):
        data, gt_mask, voxel_size = tools.load_pickle_data(f)
        data = tools.windowing(data)
        data = data[s,...]
        gt_mask = gt_mask[s,...]

        plt.figure()
        plt.subplot(121), plt.imshow(data, 'gray', interpolation='nearest'), plt.axis('off')
        # plt.subplot(122), plt.imshow(gt_mask, 'gray', interpolation='nearest')
        plt.subplot(122), plt.imshow(skiseg.mark_boundaries(data, gt_mask==1, color=(1, 0, 0), mode='thick'), interpolation='nearest'), plt.axis('off')
        plt.suptitle(f.split('/')[-1] + ', ' + str(s))
    plt.show()